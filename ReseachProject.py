import pandas as pd
import openpyxl
# set variable for first excel sheet
sheet1 = openpyxl.load_workbook('store 1.xlsx')
sheet1 = sheet1.active
cells = sheet1['A1':'B6']
# set variable for second excel sheet
sheet2 = openpyxl.load_workbook('store 2.xlsx')
sheet2 = sheet2.active
# Create a dictionary of prices for each supermarket
supermarket1 = {}
for row in sheet1:
    item = row[0].value
    price = row[1].value
    supermarket1[item] = price
supermarket2 = {}
for row in sheet2:
    item = row[0].value
    price = row[1].value
    supermarket2[item] = price
#supermarket1 = {f"{sheet1.cell(row=i, column=1).value}": sheet1.cell(row=i, column=2).value for i in range(2, sheet1.max_row + 1)}
#supermarket2 = {f"{sheet2.cell(row=i, column=1).value}": sheet2.cell(row=i, column=2).value for i in range(2, sheet2.max_row + 1)}
# Convert the dictionaries to pandas dataframes
df1 = pd.DataFrame.from_dict(supermarket1, orient='index', columns=['Supermarket 1'])
df2 = pd.DataFrame.from_dict(supermarket2, orient='index', columns=['Supermarket 2'])
# Merge the dataframes on the index (product name)
merged_df = pd.merge(df1, df2, left_index=True, right_index=True)
# Add a column to show the cheaper price for each product
merged_df['Cheaper Price'] = merged_df.apply(lambda row: ('Supermarket 1') if row['Supermarket 1'] < row['Supermarket 2'] else ('Supermarket 2'), axis=1)
# Print the merged dataframe with the cheaper price for each product
print(merged_df)